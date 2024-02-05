import easygui
import openpyxl as ex

def OpenFileEXL():
    return easygui.fileopenbox("Таблица Excel", "Открыть")
    pass
def FindPrice(str, path):
    print(454)
    wb = ex.load_workbook(path)
    sheet = wb.active
    i = 1
    print(4554)
    while (sheet.cell(i, 1).value != None):
        if (sheet.cell(i, 1).value == str):
            print(sheet.cell(i, 1).value, sheet.cell(i, 2).value,  665)
            return float(sheet.cell(i, 2).value)
        i += 1
    return 0