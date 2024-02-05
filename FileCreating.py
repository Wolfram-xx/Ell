import math

import easygui
import threading

from fpdf import FPDF
import resourses

import pandas as excel
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ImageEx
from PIL import Image, ImageDraw, ImageFont
from openpyxl.styles import Alignment, Font, Border, Side


numberDetails = []
nameDetails = []
numDetails = []
imgDetails = []
OutDiamDetails = []
InnDiamDetails = []
LengthDetails = []
CostDetails = []
NData = []
Data = []
numDet = []
nameTable = []

def cleararray():
    numberDetails.clear()
    nameDetails.clear()
    numDetails.clear()
    imgDetails.clear()
    OutDiamDetails.clear()
    InnDiamDetails.clear()
    LengthDetails.clear()
    CostDetails.clear()
    NData.clear()
    Data.clear()
    numDet.clear()
    nameTable.clear()

def sizeimg(img):
    with Image.open(img) as img_1:
        img_1.load()
    return (img_1.height * 50) // img_1.width

def createPDF(self, Path, Data):
    pdf = FPDF()
    # директория где лежат системные шрифты OS Linux
    font_dir = 'Fonts'
    # добавляем TTF-шрифт, поддерживающий кириллицу.
    pdf.add_font("Serif", style="", fname=f"{font_dir}/ofont.ru_Times New Roman.ttf", uni=True)
    pdf.add_page()
    pdf.set_font("Serif", size=10)
    # высота ячейки
    line_height = pdf.font_size * 2.5

    print(11)
    commercialOfferNumber = f"{Data[14]}"
    dateOfTheCommercialOffer = f"{Data[5]}"
    customer = f"{Data[0]}"
    executor = "ПАО «Ижнефтемаш»"
    pdf.image("icon/header.jpeg", w=150, x=30)
    pdf.set_x(20)
    pdf.multi_cell(170, 5, f"Коммерческое предложение № {commercialOfferNumber}", align='C')
    pdf.set_x(20)
    pdf.cell(85, 5, "г. Ижевск", align='L')
    pdf.multi_cell(85, 5, dateOfTheCommercialOffer, align='R')
    pdf.set_x(20)
    pdf.multi_cell(170, 5, f"Заказчик: {customer}\nИсполнитель: {executor}")
    pdf.set_x(20)
    pdf.multi_cell(170, 5,
                   "1. Наименование Услуг / Перечень оборудования (ассортимент, количество, комплектность, особые условия изготовления):")
    pdf.set_x(20)
    pdf.multi_cell(170, 5,
                   "Оказание Услуг по инженерному сопровождению спуска хвостовика с предоставлением комплекта Оборудования FRACTURA")
    pdf.set_font("Serif", size=7)
    nameLayout = f"{Data[2]} портовая Компоновка МГРП, {Data[10]}"
    y = 80
    pdf.set_xy(20, y)
    print(22)
    pdf.multi_cell(170, 5, nameLayout, 1)
    y += 5
    if Data[11] != "":
        pdf.set_xy(20, y)
        n = math.ceil(len(Data[11]) / 144)
        pdf.multi_cell(170, 5, f"{Data[11]}", 1)
        y += n*5
    pdf.set_xy(20, y)
    pdf.cell(9, 6, '№ п/п', 1, 1, "C")  # номер
    pdf.set_xy(29, y)
    pdf.cell(35, 6, 'Шифр', 1, 1, "C")  # шифр
    pdf.set_xy(64, y)
    pdf.multi_cell(60, 6, 'Наименование', 1, "L")  # наименование
    pdf.set_xy(124, y)
    pdf.multi_cell(10, 6, 'Ед.изм.', 1, "C")  # единицы
    pdf.set_xy(134, y)
    pdf.multi_cell(10, 6, 'Кол-во', 1, "C")  # кол-во
    pdf.set_xy(144, y)
    pdf.multi_cell(23, 3, 'Цена за единицу, тыс. руб. без НДС', 1, "C")  # цена
    pdf.set_xy(167, y)
    pdf.multi_cell(23, 3, 'Стоимость, тыс. руб. без НДС', 1, "C")  # стоимость
    y += 6
    print(33)
    pdf.set_xy(20, y)
    pdf.cell(170, 5, "Стоимость Оборудования", 1, 1, "C")
    DataInfTable = []
    cost_all = 0
    print(Data[12], Data[13])
    for i in range(len(Data[12])):
        print(i)
        cost = round(float(Data[13][i])/Data[12][i][3], 2)
        cost_all += float(Data[13][i])
        DataInfTable.append([str(i+1), Data[12][i][2], Data[12][i][1], "шт.", str(Data[12][i][3]), str(cost), Data[13][i]])
    y += 5
    for row in DataInfTable:
        print(row)
        n = len(row[2]) / 47
        pdf.set_xy(20, y)
        pdf.cell(9, 6, row[0], 1, 1, "C")  # номер
        pdf.set_xy(29, y)
        pdf.cell(35, 6, row[1], 1, 1, "C")  # шифр
        pdf.set_xy(64, y)
        if n > 1:
            pdf.multi_cell(60, 3, row[2], 1, "L")  # наименование
        else:
            pdf.multi_cell(60, 6, row[2], 1, "L")  # наименование
        pdf.set_xy(124, y)
        pdf.multi_cell(10, 6, row[3], 1, "C")  # единицы
        pdf.set_xy(134, y)
        pdf.multi_cell(10, 6, row[4], 1, "C")  # кол-во
        pdf.set_xy(144, y)
        pdf.multi_cell(23, 6, row[5], 1, "R")  # цена
        pdf.set_xy(167, y)
        pdf.multi_cell(23, 6, row[6], 1, "R")  # стоимость
        y += 6
    pr = "20%"
    print(cost_all)
    pdf.set_xy(20, y)
    pdf.multi_cell(124, 5, "Стоимость без НДС", "LB", "R")
    pdf.set_xy(144, y)
    pdf.multi_cell(23, 5, "", "RB", "R")
    pdf.set_xy(167, y)
    pdf.multi_cell(23, 5, str(round(cost_all, 1)), 1, "R")

    y += 5
    pdf.set_xy(20, y)
    pdf.multi_cell(124, 5, "НДС", "LB", "R")
    pdf.set_xy(144, y)
    pdf.multi_cell(23, 5, pr, "RB", "R")
    pdf.set_xy(167, y)
    nds = round((20 * cost_all) / 100, 2)
    pdf.multi_cell(23, 5, str(nds), 1, "R")

    y += 5
    pdf.set_xy(20, y)
    pdf.multi_cell(124, 5, "Стоимость с НДС", "LB", "R")
    pdf.set_xy(144, y)
    pdf.multi_cell(23, 5, "", "RB", "R")
    pdf.set_xy(167, y)
    cost_nds = nds+cost_all
    pdf.multi_cell(23, 5, str(cost_nds), 1, "R")
    print(55)
    pdf.output(Path)
def createExcel(self, Path, Data):
    projectInfo = excel.DataFrame({
        'NData': [],
        'Data': []
    })
    projectInfo.to_excel(Path)
    excel.set_option('max_colwidth', 120)
    excel.set_option('display.width', 500)
    with excel.ExcelWriter(Path) as writer:
        projectInfo.to_excel(writer, sheet_name='Схема', header=False, index=False, startcol=6, startrow=1)
    wb = load_workbook(Path)
    ws = wb.active
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20

    ws.row_dimensions[4].height = 30
    ws.row_dimensions[8].height = 30
    ws.row_dimensions[16].height = 30

    keepNumCells = ['A1:I1', 'A2:I2',
                    'B3:E3', 'B4:E4', 'B5:E5', 'B6:E6', 'J1:L2', 'J3:L3', 'K4:L4', 'J5:J6', 'K5:L6',
                    'A7:I7', 'J7:L7', 'J10:J11', 'K10:L11', 'C16:I16',
                    'C8:D8', 'E8:F8', 'K8:L8', 'K9:L9', 'K10:L10', 'K11:L11',
                    'K12:L12', 'K13:L13', 'K14:L14', 'K15:L15']
    for n in keepNumCells:
        ws.merge_cells(n)
    mediumlt_border = Border(left=Side(style='medium'),
                             top=Side(style='medium'))
    mediuml_border = Border(left=Side(style='medium'))
    mediumlb_border = Border(left=Side(style='medium'),
                             bottom=Side(style='medium'))
    mediumb_border = Border(bottom=Side(style='medium'))
    mediumtr_border = Border(right=Side(style='medium'),
                             top=Side(style='medium'))
    mediumr_border = Border(right=Side(style='medium'))
    mediumrb_border = Border(right=Side(style='medium'),
                             bottom=Side(style='medium'))
    mediumtb_border = Border(top=Side(style='medium'),
                             bottom=Side(style='medium'))
    ws["A1"] = f"{Data[0]}, скважина {Data[1]} / {Data[2]} Стадий ГРП"
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].font = Font(bold=True, size="16", name="Arial")
    ws["A1"].border = mediumlt_border
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws["A2"] = "Заканчивание скважины"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(bold=True, size="16", name="Arial")
    ws["A2"].border = mediumlb_border
    ws["J1"].border = Border(left=Side(style='medium'),
                             right=Side(style='medium'),
                             top=Side(style='medium'),
                             bottom=Side(style='medium'))
    numCellPrIn = ["A3", "A4", "A5", "A6",
                   "F3", "F4", "F5", "F6",
                   "H3", "H4", "H5", "H6",
                   "J3", "J4", "J5",
                   "A9", "A10", "A11", "A12", "A13", "A14", "A15",
                   "B8", "C8", "E8", "G8", "H8", "I8",
                   "J8", "J9", "J10", "J12", "J13", "J14", "J15",
                   "A16", "C16", "J16", "K16", "L16"]
    infCellPrIn = ["Заказчик:", "Месторождение:", "Представитель:", "Представитель:",
                   "Дата:", "Скв.", "тел:", "тел:",
                   "Сервис центр", "Тел.СЦ", "Представитель 1", "Представитель 2",
                   "Тип: ", "Диаметр открытого ствола:", "Марка стали:",
                   "КОНДУКТОР", "ЭКС. КОЛОННА", "ЭКС. КОЛОННА", "ЭКС. КОЛОННА", "ХВОСТОВИК", "ОТКР. СТВОЛ", "НКТ",
                   "Диаметр", "Толщина стенки", "Марка стали", "Тип резьбы", "Глубина от", "Глубина до",
                   "Начало набора угла:", "Глуб. по верт.:", "Пласт.давление:",
                   "Температура:", "Стол ротора", "Бур.раствор", "Раствор заканчивания:",
                   "Глубина (м)", "Описание", "Наружный диаметр", "Внутренний диаметр", "Длина"]
    for i in range(len(numCellPrIn)):
        n = numCellPrIn[i]
        ws[n] = infCellPrIn[i]
        ws[n].alignment = Alignment(horizontal="center", wrap_text=True)
        ws[n].font = Font(bold=True, size="12", name="Arial")
    ws['J4'].font = Font(bold=True, size="11", name="Arial")

    numCellValue = ["B3", "B4", "B5", "B6", "G3", "G4", "G5", "G6", "I3", "I4", "I5", "I6",
                    "K4", "K5", "K8", "K9", "K10", "K12", "K13", "K14", "K15"]
    infCellValue = [f"{Data[0]}", f"{Data[3]}", f"{Data[4]}", "", f"{Data[5]}", f"{Data[1]}", f"{Data[6]}", "",
                    "", "", "", "",
                    f"{Data[7]}", "", "", "", f"{Data[8]}", "", "", "", ""]
    """[Company, Field, Engineer, Data, Well, Contact1, Contact2,
                    ServiceCenter, ContactSerCenter, Agent1,
                    DiamOpenF, MarkSteel, Corner, VerticalDepth, ReservoirPres, Temperature, RotorTable, DrillingMud, FinishingSolution]"""
    for i in range(len(numCellValue)):
        n = numCellValue[i]
        ws[n] = infCellValue[i]
        ws[n].font = Font(size="12", name="Arial")

    ws["J3"].border = Border(left=Side(style='medium'),
                             right=Side(style='medium'))
    ws["J4"].border = mediuml_border
    ws["K4"].border = mediumr_border
    ws["J5"].border = mediumlb_border
    ws["K5"].border = mediumrb_border
    ws["J8"].border = mediuml_border
    ws["K8"].border = mediumr_border
    ws["J9"].border = mediuml_border
    ws["K9"].border = mediumr_border
    ws["J10"].border = mediuml_border
    ws["K10"].border = mediumr_border
    ws["J12"].border = mediuml_border
    ws["K12"].border = mediumr_border
    ws["J13"].border = mediuml_border
    ws["K13"].border = mediumr_border
    ws["J14"].border = mediuml_border
    ws["K14"].border = mediumr_border
    ws["J15"].border = mediumlb_border
    ws["K15"].border = mediumrb_border
    ws["J16"].border = mediumlb_border
    ws["K16"].border = mediumb_border
    ws["L16"].border = mediumrb_border
    ws["A16"].border = mediumtb_border
    ws["B16"].border = mediumtb_border
    ws["C16"].border = mediumtb_border

    ws["A7"] = "ДАННЫЕ ПО ТРУБАМ"
    ws["A7"].alignment = Alignment(horizontal="center")
    ws["A7"].font = Font(bold=True, size="12", name="Arial")
    ws["A7"].border = mediumtr_border
    ws["J7"] = "ДАННЫЕ ПО СКВАЖИНЕ"
    ws["J7"].alignment = Alignment(horizontal="center")
    ws["J7"].font = Font(bold=True, size="12", name="Arial")
    ws["J7"].border = mediumtr_border

    for i in ["C9:D9", "C10:D10", "C11:D11", "C12:D12", "C13:D13", "C14:D14", "C15:D15",
              "E9:F9", "E10:F10", "E11:F11", "E12:F12", "E13:F13", "E14:F14", "E15:F15"]:
        ws.merge_cells(i)

    details = Data[9]
    infCellDValue = []
    for detail in details:
        infCellDValue.append([detail[4], detail[2], detail[3], detail[5], detail[6], detail[7]])

    def sizeimg(img):
        with Image.open(img) as img_1:
            img_1.load()
        return (img_1.height * 50) // img_1.width
    if Data[2] >=1:
        cst = int(Data[2])
    else:
        cst = 0
    r = 25
    for i in range(len(infCellDValue)):
        img = str(infCellDValue[i][0])
        logo = ImageEx(img)
        logo.width = 50
        logo.height = sizeimg(img)
        ws.add_image(logo, "B" + str(r + 1))
        if i < len(infCellDValue)-1:
            img2 = "image/000.png"
            log = ImageEx(img2)
            log.width = 50
            log.height = sizeimg(img2)
            ws.add_image(log, "B" + str(r + 1 + round(logo.height/20.63)))
        ws["C" + str(r)] = infCellDValue[i][1]
        ws["C" + str(r)].font = Font(bold=True, size="12", name="Arial")
        ws["H" + str(r)] = infCellDValue[i][2]
        ws["H" + str(r)].font = Font(size="12", name="Arial")
        ws["J" + str(r)] = infCellDValue[i][3]
        ws["J" + str(r)].font = Font(size="12", name="Arial")
        ws["K" + str(r)] = infCellDValue[i][4]
        ws["K" + str(r)].font = Font(size="12", name="Arial")
        ws["L" + str(r)] = infCellDValue[i][5]
        ws["L" + str(r)].font = Font(size="12", name="Arial")
        if (i >= 4 and i < len(infCellDValue)-3):
            print(i)
            if i%2 == 0:
                ws["A" + str(r)] = f"Пакер {cst}"
                ws["A" + str(r)].font = Font(bold=True, size="12", name="Arial")
            else:
                ws["A" + str(r)] = f"Стадия {cst}"
                ws["A" + str(r)].font = Font(bold=True, size="12", name="Arial")
                cst -= 1
        r = r + round(logo.height/20.63) + 7
    wb.save(Path)





def NewFile(self, FileType, Data):
    print(123)
    match (FileType):
        case "PDF":
            print(1)
            input_file = easygui.filesavebox("PDF-документ", "Сохранить как", "PDFSheet.pdf", filetypes=["*.pdf"])
            print(2)
            if (input_file != None):
                createPDF(self, input_file, Data)
            cleararray()
        case "XLSX":
            input_file = easygui.filesavebox("Таблица Excel", "Сохранить как", "ExcelSheet.xlsx", filetypes=["*.xlsx"])
            if (input_file != None):
                createExcel(self, input_file, Data)
            cleararray()
